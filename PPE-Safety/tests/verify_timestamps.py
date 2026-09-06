"""
Does every safety event carry the right clock — the recording's own when the
picture shows one, the system's when it does not?

The resolver under test is `backend/app/vision/frame_clock.py`, and the
sections walk outward from it exactly the way it will fail in the field:

    1  parsing         the pinned formats, day-first, the calendar, and the
                       repo's own demo overlay with its double space
    2  the machine     lock, extrapolate, jitter, the backward jump, the
                       forward skip, coasting horizons, the looped file,
                       the sampling cadence and the seek budget — all on a
                       scripted reader and a hand-advanced clock, no OCR
    3  the store       occurred_at, timestamp_source, and both closers
                       stamping ended_at in the opened clock's own domain
                       (added with the store threading step)
    4+ the engine, the wire, the surfaces
                       (added with the integration steps; each section
                       names its step in the plan)

Run from `backend/`:

    PYTHONPATH=$PWD .venv/bin/python ../tests/verify_timestamps.py

Sections needing a live server take `--base http://127.0.0.1:8014` and are
skipped with a note until one is given.
"""

import atexit
import re
import sys
from datetime import date, datetime, time as time_of_day, timedelta, timezone
from pathlib import Path

HERE = Path(__file__).resolve().parent
REPO = HERE.parent
sys.path.insert(0, str(REPO / "backend"))

from app.vision.frame_clock import (  # noqa: E402
    ACQUIRING,
    COASTING,
    DETECT_BANDS,
    INVALID,
    LOCKED,
    NO_CLOCK,
    FrameClock,
    combine_with_anchor,
    normalize_ocr_text,
    parse_burned_timestamp,
)

failures: list[str] = []
advisories: list[str] = []


def check(name: str, ok: bool, detail: str = "") -> bool:
    print(("PASS  " if ok else "FAIL  ") + name
          + (f"  [{detail}]" if detail and not ok else ""))
    if not ok:
        failures.append(name)
    return ok


def note(name: str, ok: bool, detail: str = "") -> bool:
    print(("PASS  " if ok else "NOTE  ") + name
          + (f"  [{detail}]" if detail and not ok else ""))
    if not ok:
        advisories.append(name)
    return ok


def section(title: str) -> None:
    print()
    print(f"--- {title}")


# ----------------------------------------------------------------------
# 1 · parsing — the pinned formats, and only them
# ----------------------------------------------------------------------

section("1 · parsing — the pinned formats, day-first, the calendar")

parsed = parse_burned_timestamp("02-04-2026 18:54:29")
check("DD-MM-YYYY HH:MM:SS parses",
      parsed is not None and parsed[0] == date(2026, 4, 2)
      and parsed[1] == time_of_day(18, 54, 29),
      f"{parsed}")

check("and 02-04-2026 is the second of April — day-first, as the "
      "requirement's own example pins",
      parsed is not None and parsed[0].month == 4 and parsed[0].day == 2,
      f"{parsed}")

slashed = parse_burned_timestamp("18/04/2026 09:41:07")
check("DD/MM/YYYY HH:MM:SS parses",
      slashed is not None and slashed[0] == date(2026, 4, 18),
      f"{slashed}")

iso_style = parse_burned_timestamp("2026-04-02 18:54:29")
check("YYYY-MM-DD HH:MM:SS parses",
      iso_style is not None and iso_style[0] == date(2026, 4, 2),
      f"{iso_style}")

bare = parse_burned_timestamp("CAM01  18:54:29")
check("a bare HH:MM:SS is a clock with no date",
      bare is not None and bare[0] is None
      and bare[1] == time_of_day(18, 54, 29),
      f"{bare}")

demo = parse_burned_timestamp("06-08-2026  09:41:07")
check("the repo's own demo overlay — date, two spaces, time — parses",
      demo is not None and demo[0] == date(2026, 8, 6)
      and demo[1] == time_of_day(9, 41, 7),
      f"{demo}")

confused = parse_burned_timestamp("O2-O4-2026 18:54:29")
check("OCR's O-for-0 confusion is repaired inside digit runs",
      confused is not None and confused[0] == date(2026, 4, 2),
      f"normalised: {normalize_ocr_text('O2-O4-2026 18:54:29')!r}")

# Both measured off this engine reading this product's own overlay: the
# same strip came back with a doubled hyphen on one crop and glued to the
# camera name with an em-dash on another.
doubled = parse_burned_timestamp("¥02-04--202618:54:29")
check("a doubled hyphen and digits glued to the time still read as one "
      "clock",
      doubled is not None and doubled[0] == date(2026, 4, 2)
      and doubled[1] == time_of_day(18, 54, 29),
      f"{doubled}")

glued = parse_burned_timestamp("CAM0102-04-—202618:54:29")
check("the camera name glued on with an em-dash does not cost the date",
      glued is not None and glued[0] == date(2026, 4, 2)
      and glued[1] == time_of_day(18, 54, 29),
      f"{glued}")

check("but a time glued onto unclaimed digits, with no date to anchor "
      "the split, is refused rather than guessed",
      parse_burned_timestamp("202618:54:29") is None,
      f"{parse_burned_timestamp('202618:54:29')}")

# Measured off a real DVR export (an operator's own recording): the same
# chip came back "18:54:32", "18:54.35" and "18:54-29" on different
# frames, and refusing the last two threw away two thirds of an
# otherwise readable clock.
dotted = parse_burned_timestamp("02-04-202618:54.35")
check("the seconds colon misread as a dot is repaired inside a time",
      dotted is not None and dotted[0] == date(2026, 4, 2)
      and dotted[1] == time_of_day(18, 54, 35),
      f"normalised: {normalize_ocr_text('02-04-202618:54.35')!r}")

dashed_sec = parse_burned_timestamp("02-04-2026 18:54-29")
check("and misread as a dash",
      dashed_sec is not None
      and dashed_sec[1] == time_of_day(18, 54, 29),
      f"normalised: {normalize_ocr_text('02-04-2026 18:54-29')!r}")

check("a date's own dashes are never touched by that repair — no colon "
      "ever precedes them",
      normalize_ocr_text("02-04-2026 18:54:29") == "02-04-2026 18:54:29",
      f"{normalize_ocr_text('02-04-2026 18:54:29')!r}")

check("the DVR's timeline ruler is not a clock: hour marks alone parse "
      "to nothing",
      parse_burned_timestamp("16:00 17.00 18.00 19:00") is None,
      f"{parse_burned_timestamp('16:00 17.00 18.00 19:00')}")

check("and two ruler labels glued by a dot are not given fake seconds — "
      "the repair refuses digits that run on",
      parse_burned_timestamp("18:00.19:00") is None
      and "18:00:19" not in normalize_ocr_text("18:00.19:00"),
      f"normalised: {normalize_ocr_text('18:00.19:00')!r}")

check("the colon lost as well as glued is still refused — where the "
      "hour starts would be a guess",
      parse_burned_timestamp("02-04-20261854:36") is None,
      f"{parse_burned_timestamp('02-04-20261854:36')}")

check("a month of 13 is a failed read, not a silent month-first retry",
      parse_burned_timestamp("02-13-2026 10:00:00") is None,
      f"{parse_burned_timestamp('02-13-2026 10:00:00')}")

check("the 31st of February is a failed read",
      parse_burned_timestamp("31-02-2026 10:00:00") is None,
      f"{parse_burned_timestamp('31-02-2026 10:00:00')}")

check("a date with no time on it is not a clock",
      parse_burned_timestamp("02-04-2026") is None,
      f"{parse_burned_timestamp('02-04-2026')}")

check("free text with no clock in it parses to nothing",
      parse_burned_timestamp("ENTRANCE CAM 3 — recording") is None, "")

check("a year outside 2000-2099 is refused",
      parse_burned_timestamp("02-04-1999 10:00:00") is None
      and parse_burned_timestamp("02-04-2101 10:00:00") is None, "")

check("an impossible time is refused",
      parse_burned_timestamp("02-04-2026 25:00:00") is None, "")

anchored = combine_with_anchor(date(2026, 4, 2), time_of_day(18, 54, 29),
                               datetime(2030, 1, 1))
check("a read with its own date ignores the expectation entirely",
      anchored == datetime(2026, 4, 2, 18, 54, 29), f"{anchored}")

rollover = combine_with_anchor(
    None, time_of_day(0, 0, 3), datetime(2026, 8, 18, 0, 0, 3)
)
check("midnight rollover: 00:00:03 expected just after midnight lands on "
      "the new day",
      rollover == datetime(2026, 8, 18, 0, 0, 3), f"{rollover}")

behind = combine_with_anchor(
    None, time_of_day(23, 59, 57), datetime(2026, 8, 18, 0, 0, 2)
)
check("and 23:59:57 read moments after midnight lands on the day before",
      behind == datetime(2026, 8, 17, 23, 59, 57), f"{behind}")

plain = combine_with_anchor(
    None, time_of_day(12, 0, 5), datetime(2026, 8, 17, 12, 0, 0)
)
check("an ordinary mid-day read stays on its own day",
      plain == datetime(2026, 8, 17, 12, 0, 5), f"{plain}")


# ----------------------------------------------------------------------
# 2 · the state machine — scripted reader, hand-advanced clock
# ----------------------------------------------------------------------

section("2 · the machine — lock, extrapolate, invalidate, coast, loop")


class Bench:
    """
    One FrameClock wired to a scripted reader and a settable clock.

    With no ROI the hunt scans both overlay bands, so one *attempt* is two
    reader calls; `calls` counts attempts — first-band calls — because the
    cadence and the budget are per attempt, not per band.
    """

    def __init__(self):
        self.mono = 1000.0
        self.script: list = []
        self.calls = 0
        self.clock = FrameClock(
            source_key="bench",
            ocr=self._read,
            mono=lambda: self.mono,
        )

    def _read(self, frame, region, detect):
        if region == DETECT_BANDS[0]:
            self.calls += 1
        if not self.script:
            return []
        return [(self.script.pop(0), None)]

    def show(self, text_or_none, pos, advance=2.0):
        """Advance time, script one reading (or nothing), feed a frame."""
        self.mono += advance
        self.script = [] if text_or_none is None else [text_or_none]
        self.clock.observe_frame(None, pos)


def burned(moment: datetime) -> str:
    return moment.strftime("%d-%m-%Y %H:%M:%S")


T0 = datetime(2026, 4, 2, 18, 54, 29)

bench = Bench()
bench.show(burned(T0), pos=10.0)
check("one valid read is acquiring, not locked — a single read is one "
      "misread digit from a wrong clock",
      bench.clock.state == ACQUIRING and bench.clock.resolve(10.0) is None,
      f"state={bench.clock.state}")

bench.show(burned(T0.replace(second=31)), pos=12.0)
stamp = bench.clock.resolve(12.0)
check("a second read two seconds later, showing two seconds more, locks",
      bench.clock.state == LOCKED and stamp is not None,
      f"state={bench.clock.state}, stamp={stamp}")

check("and the stamp is the burned time, not the wall's",
      stamp is not None and stamp.naive == "2026-04-02 18:54:31",
      f"{stamp}")

NOW_SHAPE = re.compile(r"^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}\+00:00$")
check("the stamp's iso is byte-shaped exactly like the store's _now()",
      stamp is not None and NOW_SHAPE.match(stamp.iso) is not None
      and NOW_SHAPE.match(
          datetime.now(timezone.utc).isoformat(timespec="seconds")
      ) is not None,
      f"{stamp.iso if stamp else None}")

half = bench.clock.resolve(12.5)
check("between reads the clock extrapolates by the video's own position",
      half is not None and half.naive == "2026-04-02 18:54:31",
      f"resolve(12.5) -> {half.naive if half else None} "
      f"(sub-second truncated)")

later = bench.clock.resolve(17.0)
check("five seconds of position is five seconds of clock",
      later is not None and later.naive == "2026-04-02 18:54:36",
      f"{later.naive if later else None}")

bench.show(burned(T0.replace(second=32)), pos=14.0)
check("a read half a second slow re-anchors rather than alarming — OCR "
      "latency and one-second display granularity both land late",
      bench.clock.state == LOCKED
      and bench.clock.resolve(14.0).naive == "2026-04-02 18:54:32",
      f"state={bench.clock.state}")

bench.show(burned(T0.replace(second=24)), pos=16.0)
check("a clock that jumps backward is invalid, and the system answers",
      bench.clock.state == INVALID and bench.clock.resolve(16.0) is None,
      f"state={bench.clock.state}")

bench.show(burned(T0.replace(second=40)), pos=18.0)
bench.show(burned(T0.replace(second=42)), pos=20.0)
relocked = bench.clock.resolve(20.0)
check("two reads that agree with each other re-lock after an invalidation",
      bench.clock.state == LOCKED and relocked is not None
      and relocked.naive == "2026-04-02 18:54:42",
      f"state={bench.clock.state}, {relocked}")

# Forward skip: stitched footage jumps ahead. One read is a candidate; the
# old anchor keeps answering until a second read agrees.
JUMP = datetime(2026, 4, 2, 19, 30, 0)
bench.show(burned(JUMP), pos=22.0)
held = bench.clock.resolve(22.0)
check("a forward skip is held as a candidate while the old timeline keeps "
      "answering",
      held is not None and held.naive == "2026-04-02 18:54:44",
      f"{held.naive if held else None}, state={bench.clock.state}")

bench.show(burned(JUMP.replace(second=2)), pos=24.0)
confirmed = bench.clock.resolve(24.0)
check("a second read on the skipped timeline confirms it",
      bench.clock.state == LOCKED and confirmed is not None
      and confirmed.naive == "2026-04-02 19:30:02",
      f"{confirmed.naive if confirmed else None}")

# A contradicted skip: one stray read ahead, then the true clock again.
bench.show(burned(datetime(2026, 4, 2, 21, 0, 0)), pos=26.0)
bench.show(burned(JUMP.replace(second=6)), pos=28.0)
steady = bench.clock.resolve(28.0)
check("one stray read ahead does not move the clock when the next read "
      "contradicts it",
      steady is not None and steady.naive == "2026-04-02 19:30:06",
      f"{steady.naive if steady else None}, state={bench.clock.state}")

# Coasting: the overlay disappears; position keeps the arithmetic exact.
bench.script = []
bench.show(None, pos=30.0)
coasting = bench.clock.resolve(30.0)
check("an unreadable sample while anchored coasts on the video position",
      bench.clock.state == COASTING and coasting is not None
      and coasting.naive == "2026-04-02 19:30:08",
      f"state={bench.clock.state}")

far = bench.clock.resolve(28.0 + 89.0)
check("ninety source-seconds is still inside the position leash",
      far is not None, f"{far}")

gone = bench.clock.resolve(28.0 + 95.0)
check("beyond the leash the honest answer is the system clock, and the "
      "hunt re-arms",
      gone is None and bench.clock.state == NO_CLOCK,
      f"state={bench.clock.state}")

# Wall-anchored: a source with no position gets the short leash.
wall = Bench()
W0 = datetime(2026, 8, 17, 9, 0, 0)
wall.show(burned(W0), pos=None)
wall.show(burned(W0.replace(second=2)), pos=None, advance=2.0)
check("a positionless source locks by wall time",
      wall.clock.state == LOCKED and wall.clock.resolve(None) is not None,
      f"state={wall.clock.state}")

wall.mono += 10.0
check("ten wall-seconds of coasting still answers cctv",
      wall.clock.resolve(None) is not None, "")

wall.mono += 6.0
check("sixteen wall-seconds is past the wall leash — system, and re-arm",
      wall.clock.resolve(None) is None and wall.clock.state == NO_CLOCK,
      f"state={wall.clock.state}")

# The looped file: position falls back to the start; the burned clock
# legitimately restarts with it.
loop = Bench()
loop.show(burned(T0), pos=100.0)
loop.show(burned(T0.replace(second=31)), pos=102.0)
check("a locked clock on a playing file",
      loop.clock.state == LOCKED, f"state={loop.clock.state}")

loop.show(burned(T0), pos=0.5)
check("the file looping is a new segment to re-acquire, never an "
      "invalidation",
      loop.clock.state == ACQUIRING and loop.clock.resolve(0.5) is None,
      f"state={loop.clock.state}")

loop.show(burned(T0.replace(second=31)), pos=2.5)
check("and the second lap locks again like any fresh playback",
      loop.clock.state == LOCKED
      and loop.clock.resolve(2.5).naive == "2026-04-02 18:54:31",
      f"state={loop.clock.state}")

# Cadence: a locked clock reads once per SAMPLE_SECONDS of source time,
# and a non-due frame never reaches the reader.
pace = Bench()
pace.show(burned(T0), pos=10.0)
pace.show(burned(T0.replace(second=31)), pos=12.0)
before = pace.calls
pace.script = [burned(T0.replace(second=32))]
pace.mono += 0.2
pace.clock.observe_frame(None, 12.5)
check("a frame half a source-second after the last read never reaches "
      "the reader",
      pace.calls == before, f"calls {before} -> {pace.calls}")

pace.mono += 0.2
pace.clock.observe_frame(None, 14.1)
check("and the next due frame does",
      pace.calls == before + 1, f"calls {before} -> {pace.calls}")

# The seek budget: a source with no clock in it stops paying quickly.
quiet = Bench()
for _ in range(12):
    quiet.show(None, pos=None, advance=0.8)
spent = quiet.calls
quiet.show(None, pos=None, advance=0.8)
check("twelve empty hunts exhaust the seek budget and the thirteenth "
      "frame is free",
      spent == 12 and quiet.calls == 12,
      f"calls={quiet.calls} after budget")

quiet.show(None, pos=None, advance=61.0)
check("a dormant clock re-probes about once a minute, not never",
      quiet.calls == 13, f"calls={quiet.calls}")

check("and a quiet source answers the system clock throughout",
      quiet.clock.resolve(None) is None
      and quiet.clock.state == NO_CLOCK,
      f"state={quiet.clock.state}")


# ----------------------------------------------------------------------
# 3 · the store — occurred_at, timestamp_source, and same-domain closes
# ----------------------------------------------------------------------

section("3 · the store — the resolved clock lands, and the close stays in "
        "its domain")

import tempfile  # noqa: E402
import time as _time  # noqa: E402

import numpy as np  # noqa: E402

import app.events.store as store_module  # noqa: E402
from app.events.store import EventStore  # noqa: E402
from app.vision.frame_clock import ResolvedStamp  # noqa: E402

WORKDIR = Path(tempfile.mkdtemp(prefix="verify_timestamps_"))

BURNED = ResolvedStamp(
    iso="2026-04-02T18:54:29+00:00",
    epoch=datetime(2026, 4, 2, 18, 54, 29, tzinfo=timezone.utc).timestamp(),
    raw="02-04-2026 18:54:29",
    naive="2026-04-02 18:54:29",
)

FINDING = [{
    "key": "intrusion-zone-1",
    "severity": "high",
    "summary": "Someone entered the restricted area",
    "details": {"zone_id": 1},
}]

# A throwaway observe first: the store lazily imports the camera register
# (and behind it the model stack) inside the first insert, and that cold
# import must not be measured as event duration below.
EventStore(path=WORKDIR / "warmup.db").observe(
    "restricted-zone", FINDING, None
)

cctv_store = EventStore(path=WORKDIR / "cctv.db")
opened_mono = _time.monotonic()
cctv_store.observe("restricted-zone", FINDING, None, resolved=BURNED)
row = cctv_store.list(limit=1)["events"][0]

check("an event opened with a resolved stamp occurs at the burned time, "
      "not the wall's",
      row["occurred_at"] == "2026-04-02T18:54:29+00:00",
      f"{row['occurred_at']}")

check("and its details say where the clock came from",
      row["details"].get("timestamp_source") == "cctv"
      and row["details"].get("cctv_timestamp") == "2026-04-02 18:54:29"
      and row["details"].get("cctv_raw") == "02-04-2026 18:54:29",
      f"{row['details']}")

check("the module's own detail fields still ride along",
      row["details"].get("zone_id") == 1, f"{row['details']}")

# The ordinary closer: the finding stops appearing. The open entry's
# last_seen is aged by hand so the suite does not sleep out the resolve
# grace — the path exercised is still observe() -> _close_absent.
for entry in cctv_store._open.values():
    entry["last_seen"] -= 10.0
cctv_store.observe("restricted-zone", [], None)
elapsed = _time.monotonic() - opened_mono
closed = cctv_store.list(limit=1)["events"][0]

check("the ordinary closer ends it", closed["ended_at"] is not None,
      f"{closed}")

lasted = (
    datetime.fromisoformat(closed["ended_at"])
    - datetime.fromisoformat(closed["occurred_at"])
).total_seconds()
check("and the April event lasted exactly the seconds it really lasted on "
      "its own clock — not the four months to the August replay",
      0.0 <= lasted <= elapsed + 2.0,
      f"occurred {closed['occurred_at']}, ended {closed['ended_at']} — "
      f"{lasted:.0f}s recorded against {elapsed:.1f}s measured")

# The no-frame closer: the camera goes away with the event still open.
cctv_store.observe("restricted-zone", FINDING, None, resolved=BURNED)
cctv_store.forget_open("restricted-zone")
dropped = cctv_store.list(limit=1)["events"][0]
drop_lasted = (
    datetime.fromisoformat(dropped["ended_at"])
    - datetime.fromisoformat(dropped["occurred_at"])
).total_seconds()
check("the source-gone closer also ends in the opened clock's domain — it "
      "has no resolver to ask and needs none",
      dropped["ended_at"] is not None and 0.0 <= drop_lasted < 30.0,
      f"occurred {dropped['occurred_at']}, ended {dropped['ended_at']}")

# The default path: no resolver anywhere near the call.
plain_store = EventStore(path=WORKDIR / "plain.db")
before_wall = datetime.now(timezone.utc)
plain_store.observe("restricted-zone", FINDING, None)
after_wall = datetime.now(timezone.utc)
plain_row = plain_store.list(limit=1)["events"][0]
plain_when = datetime.fromisoformat(plain_row["occurred_at"])

check("an event with no resolved stamp occurs at the system clock, as "
      "every event always has",
      before_wall - timedelta(seconds=1)
      <= plain_when
      <= after_wall + timedelta(seconds=1),
      f"{plain_row['occurred_at']} vs wall {before_wall.isoformat()}")

check("and its details name the system as the source",
      plain_row["details"].get("timestamp_source") == "system",
      f"{plain_row['details']}")

check("with no cctv keys invented for a clock that was never read",
      "cctv_timestamp" not in plain_row["details"]
      and "cctv_raw" not in plain_row["details"],
      f"{sorted(plain_row['details'])}")

# The late-locking clock: an event opens in the hunt seconds on the
# system clock, and the moment the recording's clock locks it is
# re-stamped to when it began on the recording's own timeline — with the
# original wall receipt still in server_timestamp.
late_store = EventStore(path=WORKDIR / "late.db")
late_store.observe("restricted-zone", FINDING, None)
before_restamp = late_store.list(limit=1)["events"][0]
late_store.restamp_open("restricted-zone", BURNED)
after_restamp = late_store.list(limit=1)["events"][0]

check("an event opened before the clock locked is re-stamped onto the "
      "recording's timeline",
      after_restamp["occurred_at"].startswith("2026-04-02T18:54")
      and after_restamp["details"].get("timestamp_source") == "cctv",
      f"{before_restamp['occurred_at']} -> {after_restamp['occurred_at']} "
      f"/ {after_restamp['details'].get('timestamp_source')}")

check("with the original wall receipt still in its details",
      after_restamp["details"].get("server_timestamp")
      == before_restamp["details"].get("server_timestamp"),
      f"{after_restamp['details'].get('server_timestamp')}")

late_store.forget_open("restricted-zone")
restamped_closed = late_store.list(limit=1)["events"][0]
restamp_lasted = (
    datetime.fromisoformat(restamped_closed["ended_at"])
    - datetime.fromisoformat(restamped_closed["occurred_at"])
).total_seconds()
check("and a re-stamped event still closes in the recording's domain",
      0.0 <= restamp_lasted < 30.0,
      f"lasted {restamp_lasted:.0f}s")

already = EventStore(path=WORKDIR / "already.db")
already.observe("restricted-zone", FINDING, None, resolved=BURNED)
already.restamp_open("restricted-zone", ResolvedStamp(
    iso="2026-06-01T00:00:00+00:00",
    epoch=datetime(2026, 6, 1, tzinfo=timezone.utc).timestamp(),
    raw="x", naive="2026-06-01 00:00:00",
))
untouched = already.list(limit=1)["events"][0]
check("an event that already carries a recording's clock is never "
      "re-stamped",
      untouched["occurred_at"] == "2026-04-02T18:54:29+00:00",
      f"{untouched['occurred_at']}")

# The snapshot filename is built from the shaped ISO, never the raw OCR
# text — measured with the evidence directory pointed at the suite's own
# scratch so nothing lands in the product's storage.
_real_evidence = store_module.EVIDENCE_DIR
store_module.EVIDENCE_DIR = WORKDIR / "evidence"
try:
    snap_store = EventStore(path=WORKDIR / "snap.db")
    snap_store.observe(
        "restricted-zone",
        FINDING,
        np.full((60, 80, 3), 128, dtype=np.uint8),
        resolved=BURNED,
    )
    snap_row = snap_store.list(limit=1)["events"][0]
    snap_name = snap_row.get("snapshot") or ""
finally:
    store_module.EVIDENCE_DIR = _real_evidence

check("the evidence filename is built from the shaped stamp — no colons, "
      "no slashes, no raw OCR text",
      bool(snap_name)
      and ":" not in snap_name and "/" not in snap_name
      and "02-04-2026 18:54:29" not in snap_name
      and snap_name.startswith("restricted-zone_2026-04-02T18-54-29"),
      f"{snap_name!r}")


# ----------------------------------------------------------------------
# 3b · the camera-clock verdict — the machine's word, the register's
#      lifecycle, and the record every event carries
# ----------------------------------------------------------------------

section("3b · the camera-clock verdict — one warning, a lifecycle, and "
        "never a verdict from one bad frame")

from app.camera.registry import CameraRegistry  # noqa: E402
from app.vision.frame_clock import (  # noqa: E402
    CLOCK_CHECKING,
    CLOCK_INVALID,
    CLOCK_UNAVAILABLE,
    CLOCK_VALID,
)

fresh = Bench()
check("a source that has not been probed yet is checking, not condemned",
      fresh.clock.clock_status() == CLOCK_CHECKING,
      f"{fresh.clock.clock_status()}")

fresh.show(None, pos=None, advance=0.8)
check("one failed read decides nothing — TEST 3's rule at frame one",
      fresh.clock.clock_status() == CLOCK_CHECKING,
      f"{fresh.clock.clock_status()}")

for _ in range(11):
    fresh.show(None, pos=None, advance=0.8)
check("the verdict becomes unavailable only after the whole check window "
      "is spent",
      fresh.clock.clock_status() == CLOCK_UNAVAILABLE,
      f"{fresh.clock.clock_status()} after the seek budget")

ticking = Bench()
ticking.show(burned(T0), pos=10.0)
ticking.show(burned(T0.replace(second=31)), pos=12.0)
check("a locked clock is valid", ticking.clock.clock_status() == CLOCK_VALID,
      f"{ticking.clock.clock_status()}")

ticking.show(None, pos=14.0)
check("and a transient miss leaves it valid — coasting is not a failure "
      "(TEST 3)",
      ticking.clock.clock_status() == CLOCK_VALID,
      f"state={ticking.clock.state}")

ticking.show(burned(T0.replace(second=20)), pos=16.0)
check("a backward jump is pronounced invalid, in the verdict too",
      ticking.clock.clock_status() == CLOCK_INVALID,
      f"{ticking.clock.clock_status()}")

# The register's lifecycle: one warning, resolved by a valid verdict,
# revalidated — never assumed — across a restart (TESTs 4, 5, 6).
reg = CameraRegistry(path=WORKDIR / "clock_registry.json")
reg.register("cam-clock-1", "Weldbay-1", "Laser Area")

reg.report_clock("cam-clock-1", "unavailable")
row = reg.get("cam-clock-1")
changes = [e for e in reg.log_entries(50)
           if e["event"] == "CAMERA_CLOCK_STATUS_CHANGED"]
check("an unavailable verdict marks the camera and creates one warning",
      row["camera_clock_status"] == "unavailable"
      and row["clock_warning_active"] is True
      and row["clock_warning_created_at"] is not None
      and len(changes) == 1,
      f"{row['camera_clock_status']} warning={row['clock_warning_active']} "
      f"changes={len(changes)}")

reg.report_clock("cam-clock-1", "unavailable")
reg.report_clock("cam-clock-1", "unavailable")
changes = [e for e in reg.log_entries(50)
           if e["event"] == "CAMERA_CLOCK_STATUS_CHANGED"]
check("repeating the same verdict never repeats the warning — one active "
      "warning per camera, not one per frame",
      len(changes) == 1, f"changes={len(changes)}")

reg.report_clock("cam-clock-1", "valid", last_read="02-04-2026 18:54:29")
row = reg.get("cam-clock-1")
resolved_log = [e for e in reg.log_entries(50)
                if e["event"] == "CAMERA_CLOCK_RESOLVED"]
check("a valid verdict resolves the warning, with its own log line and a "
      "validation timestamp (TEST 5)",
      row["camera_clock_status"] == "valid"
      and row["clock_warning_active"] is False
      and row["clock_last_validated_at"] is not None
      and row["last_camera_timestamp"] == "02-04-2026 18:54:29"
      and len(resolved_log) == 1,
      f"{row}")

check("the change log names both ends of every transition",
      all(e.get("previous") and e.get("new") and e.get("reason")
          for e in reg.log_entries(50)
          if e["event"] == "CAMERA_CLOCK_STATUS_CHANGED"),
      f"{[e for e in reg.log_entries(50) if e['event'] == 'CAMERA_CLOCK_STATUS_CHANGED']}")

reg.report_clock("cam-clock-ghost", "unavailable")
check("a verdict for an unregistered camera is dropped — there is no "
      "record to hang a warning on",
      reg.get("cam-clock-ghost") is None, "")

reloaded = CameraRegistry(path=WORKDIR / "clock_registry.json")
row = reloaded.get("cam-clock-1")
check("a restart revalidates rather than believes: the verdict resets to "
      "unknown, the warning clears, the history survives (TEST 6)",
      row["camera_clock_status"] == "unknown"
      and row["clock_warning_active"] is False
      and row["clock_last_validated_at"] is not None,
      f"{row}")

check("the register maps server sources to their cameras and refuses "
      "footage that is not a camera",
      reloaded.camera_for_source("cam-clock-1") == "cam-clock-1"
      and reloaded.camera_for_source("storage/uploads/x.mp4") is None,
      "")

local_reg = CameraRegistry(path=WORKDIR / "clock_registry2.json")
local_reg.register("local:0", "Bench cam", "Test rig")
check("and a local device index matches through its local: spelling",
      local_reg.camera_for_source(0) == "local:0"
      and local_reg.camera_for_source("local:0") == "local:0",
      f"{local_reg.camera_for_source(0)}")

# Every event carries the verdict, apart from the timestamp source — the
# system stamping an event must never read as the camera clock being fine.
verdict_store = EventStore(path=WORKDIR / "verdict.db")
verdict_store.observe("restricted-zone", FINDING, None,
                      clock_status="unavailable")
verdict_row = verdict_store.list(limit=1)["events"][0]
check("an event from a clockless camera says so beside its system stamp",
      verdict_row["details"].get("timestamp_source") == "system"
      and verdict_row["details"].get("camera_clock_status") == "unavailable",
      f"{verdict_row['details']}")

verdict_store.forget_open("restricted-zone")
verdict_store.observe("restricted-zone", FINDING, None)
legacy_row = verdict_store.list(limit=1)["events"][0]
check("a caller that predates the verdict records unknown, never a claim",
      legacy_row["details"].get("camera_clock_status") == "unknown",
      f"{legacy_row['details']}")


# ----------------------------------------------------------------------
# 4 · the engine — offline, on the overlay the product actually burns
# ----------------------------------------------------------------------

section("4 · the OCR engine — offline, reading the demo overlay")

import cv2  # noqa: E402

from app.vision.frame_clock import _get_engine, _engine_read  # noqa: E402

PEOPLE = cv2.imread(str(HERE / "fixtures" / "people_cctv.png"))
check("the people fixture is present",
      PEOPLE is not None and PEOPLE.shape[:2] == (480, 640),
      f"{None if PEOPLE is None else PEOPLE.shape}")


def overlay(frame, at):
    """The demo script's own overlay: black band, Hershey text, two spaces."""
    out = frame.copy()
    cv2.rectangle(out, (0, 0), (out.shape[1], 26), (0, 0, 0), -1)
    cv2.putText(out, "CAM 01", (6, 19),
                cv2.FONT_HERSHEY_SIMPLEX, 0.55, (255, 255, 255), 1)
    cv2.putText(out, at.strftime("%d-%m-%Y  %H:%M:%S"), (330, 19),
                cv2.FONT_HERSHEY_SIMPLEX, 0.55, (255, 255, 255), 1)
    return out


engine = _get_engine()

if engine is None:
    note("the OCR engine is installed", False,
         "rapidocr_onnxruntime is absent — the latch answers the system "
         "clock, and the sections that need reading are skipped")
else:
    check("the engine loads offline — its models ship inside the wheel",
          True, "")

    rows = _engine_read(overlay(PEOPLE, T0), (0.0, 0.0, 1.0, 0.14), True)
    joined = " ".join(str(row[0]) for row in rows)
    read_back = parse_burned_timestamp(joined)

    check("it reads the burned overlay off the top band",
          read_back is not None
          and read_back[0] == date(2026, 4, 2)
          and read_back[1] == time_of_day(18, 54, 29),
          f"rows {[row[0] for row in rows]!r} -> {read_back}")

    check("and the date survives the detector splitting it from the time — "
          "the two arrive as separate boxes and are parsed joined",
          read_back is not None and read_back[0] is not None,
          f"{read_back}")


# ----------------------------------------------------------------------
# 5-8 · the wire, the photos, the server path, the surfaces — live
# ----------------------------------------------------------------------

BASE = None
for index, arg in enumerate(sys.argv):
    if arg == "--base" and index + 1 < len(sys.argv):
        BASE = sys.argv[index + 1]

if BASE is None or engine is None:
    section("5-8 · the live sections")
    note("the live sections ran", False,
         "needs --base http://127.0.0.1:8014 and the OCR engine — "
         "run again with a fresh server to measure the wire, the photos, "
         "the server path and the surfaces")
else:
    import asyncio  # noqa: E402
    import io as _io  # noqa: E402
    import struct  # noqa: E402

    import requests  # noqa: E402
    import websockets  # noqa: E402

    WS_BASE = BASE.replace("http", "ws", 1)

    #: The foot-silhouette zone from the phase-3 record: the walking
    #: worker's feet are wholly inside it, so every frame of the fixture
    #: is an intrusion — a deterministic event source. Point dicts, the
    #: shape the config verb has always taken.
    ZONE = [
        {"x": 130, "y": 405}, {"x": 225, "y": 405},
        {"x": 225, "y": 443}, {"x": 130, "y": 443},
    ]

    def jpeg(frame):
        ok, buffer = cv2.imencode(".jpg", frame)
        assert ok
        return buffer.tobytes()

    def enveloped(payload, pos):
        head = b"VTS1" + bytes([1]) + struct.pack(">d", pos)
        return head + payload

    def zone_events(limit=10):
        reply = requests.get(
            f"{BASE}/api/events",
            params={"module": "restricted-zone", "days": 365,
                    "limit": limit},
            timeout=10,
        ).json()
        return reply["data"]["events"]

    def zone_latest():
        """This run's newest zone event — by insertion id, not by clock.

        Ordering by occurred_at would hand back whichever historical row
        carries the furthest-future stamp; ids only ever count up."""
        rows = zone_events(limit=200)
        return max(rows, key=lambda event: event["id"]) if rows else None

    def _clean_up():
        """
        Put the product's config back, however this suite is leaving.

        Idempotent and silent: it runs at the end of a good run and again
        on the way out of a bad one, and a server that has already gone
        away is not something a teardown should shout about.
        """
        for payload in ({"polygon": []}, None, {"polygon": []}):
            try:
                if payload is None:
                    requests.post(f"{BASE}/camera/stop", timeout=15)
                else:
                    requests.post(f"{BASE}/api/restricted-zone/config",
                                  json=payload, timeout=10)
            except Exception:  # noqa: BLE001
                pass
        try:
            requests.post(f"{BASE}/api/timestamp-clock/config",
                          json={"source": "browser", "box": None}, timeout=10)
        except Exception:  # noqa: BLE001
            pass

    atexit.register(_clean_up)

    def wait_closed(event_id, deadline=12.0):
        """
        The event, once it has ended — found by id, not by recency.

        A row stamped from footage sits on the day the recording was made,
        so it sorts *below* everything watched live today: scanning the
        newest handful missed the very rows this suite creates, and read
        as "it never closed" when the closing was fine. The window is
        wide and the match is on the id.
        """
        start = _time.monotonic()
        while _time.monotonic() - start < deadline:
            for event in zone_events(limit=200):
                if event["id"] == event_id and event["ended_at"]:
                    return event
            _time.sleep(0.5)
        return None

    async def push(url, frames, pause=0.15):
        """Send frames (payload bytes) and return the last JSON reply."""
        last = None
        async with websockets.connect(url, max_size=None) as ws:
            for payload in frames:
                await ws.send(payload)
                last = json.loads(await ws.recv())
                if "error" in last:
                    raise RuntimeError(str(last))
                await asyncio.sleep(pause)
        return last

    import json  # noqa: E402

    section("5 · the wire — a replayed recording stamps its own clock")

    marked = requests.post(
        f"{BASE}/api/restricted-zone/config",
        json={"polygon": ZONE, "frame_width": 640, "frame_height": 480},
        timeout=10,
    ).json()
    check("the intrusion zone is marked — the deterministic event source "
          "everything below stands on",
          marked.get("success") is True, f"{marked}")

    # Video A: the recording's clock starts months back and ticks with
    # its own position.
    frames_a = [
        enveloped(jpeg(overlay(PEOPLE, T0 + timedelta(seconds=int(i * 0.4)))),
                  i * 0.4)
        for i in range(30)
    ]
    last = asyncio.run(
        push(f"{WS_BASE}/api/restricted-zone/ws?overlay=json", frames_a)
    )

    check("the module saw the intrusion the fixture stages",
          bool(last and last.get("alert")), f"{last}")

    check("TEST 1 · a camera with a readable burned clock reports a valid "
          "camera clock over the wire, and shows no warning",
          last.get("camera_clock") == "valid",
          f"camera_clock={last.get('camera_clock')}")

    opened = zone_latest()

    check("the event occurred at the recording's own clock — April, not "
          "the wall's August",
          opened["occurred_at"].startswith("2026-04-02T18:54"),
          f"{opened['occurred_at']}")

    check("and names the footage as its clock",
          opened["details"].get("timestamp_source") == "cctv"
          and str(opened["details"].get("cctv_timestamp", "")).startswith(
              "2026-04-02 18:54")
          and opened["details"].get("camera_clock_status") == "valid",
          f"{opened['details']}")

    closed = wait_closed(opened["id"])
    lasted = (
        (datetime.fromisoformat(closed["ended_at"])
         - datetime.fromisoformat(closed["occurred_at"])).total_seconds()
        if closed else None
    )
    check("the socket dropping closed it in the recording's own domain — "
          "seconds, not the months to the replay",
          closed is not None and 0.0 <= lasted <= 60.0,
          f"lasted {lasted}s ({closed})")

    # Video B: no overlay at all, pushed the way every old client pushes —
    # bare JPEGs, no envelope. The default path, byte for byte — and TEST 2:
    # kept feeding until the clock's whole check window is spent and the
    # verdict lands, detection running the entire time.
    async def push_until_verdict(url, wanted, deadline=60.0):
        start = _time.monotonic()
        payload = jpeg(PEOPLE)
        last = None
        async with websockets.connect(url, max_size=None) as ws:
            while _time.monotonic() - start < deadline:
                await ws.send(payload)
                last = json.loads(await ws.recv())
                if "error" in last:
                    raise RuntimeError(str(last))
                if last.get("camera_clock") == wanted:
                    return last
                await asyncio.sleep(0.2)
        return last

    verdict_reply = asyncio.run(push_until_verdict(
        f"{WS_BASE}/api/restricted-zone/ws?overlay=json", "unavailable"
    ))

    check("TEST 2 · a source with no burned clock is pronounced "
          "unavailable after the check window — with detection replying "
          "on every frame throughout",
          verdict_reply is not None
          and verdict_reply.get("camera_clock") == "unavailable"
          and "alert" in verdict_reply,
          f"camera_clock={verdict_reply.get('camera_clock') if verdict_reply else None}")

    plain = zone_latest()
    plain_when = datetime.fromisoformat(plain["occurred_at"])
    wall = datetime.now(timezone.utc)

    check("a recording with no burned clock, on an old client's bare "
          "frames, is stamped by the system exactly as ever",
          plain["details"].get("timestamp_source") == "system"
          and abs((wall - plain_when).total_seconds()) < 120.0,
          f"{plain['occurred_at']} / {plain['details']}")

    check("with no cctv keys invented",
          "cctv_timestamp" not in plain["details"],
          f"{sorted(plain['details'])}")

    wait_closed(plain["id"])

    # A mark saved mid-session must reach the clock that is running,
    # not the next one — the defect a real operator hit: they re-marked
    # the box while their recording streamed and watched the warning
    # stand, because only a fresh socket ever read the store. The frames
    # here are an empty scene, so this opens no zone event of its own.
    blank = jpeg(np.full((480, 640, 3), 128, dtype=np.uint8))
    status_url = f"{BASE}/api/timestamp-clock/status"
    config_url = f"{BASE}/api/timestamp-clock/config"

    def browser_clocks():
        reply = requests.get(status_url, timeout=10).json()
        return [entry for entry in reply["data"]["clocks"]
                if entry.get("source") == "browser"]

    async def mark_mid_session():
        url = f"{WS_BASE}/api/restricted-zone/ws?overlay=json"
        out = {}
        async with websockets.connect(url, max_size=None) as ws:
            await ws.send(blank)
            json.loads(await ws.recv())

            out["listed"] = browser_clocks()

            box = [0.62, 0.88, 0.79, 0.96]
            out["saved"] = requests.post(
                config_url, json={"source": "browser", "box": box},
                timeout=10,
            ).json()
            out["marked"] = browser_clocks()

            out["cleared_reply"] = requests.post(
                config_url, json={"source": "browser", "box": None},
                timeout=10,
            ).json()
            out["cleared"] = browser_clocks()

            await ws.send(blank)
            out["still_answering"] = json.loads(await ws.recv())
        return out

    live = asyncio.run(mark_mid_session())

    check("TEST 7 · the status route lists a browser session's clock "
          "while its socket runs — no more debugging an empty list",
          len(live["listed"]) == 1, f"{live['listed']}")

    check("a mark saved mid-session re-arms that running clock on the "
          "spot — its reading region is the posted box, no restart "
          "involved",
          live["saved"].get("success") is True
          and len(live["marked"]) == 1
          and live["marked"][0].get("roi") == [0.62, 0.88, 0.79, 0.96],
          f"{live['marked']}")

    check("clearing the mark mid-session hands the same clock back to "
          "auto-detection",
          len(live["cleared"]) == 1
          and live["cleared"][0].get("roi") is None,
          f"{live['cleared']}")

    check("and the socket answered frames throughout the whole exchange",
          "error" not in live["still_answering"],
          f"{live['still_answering'].get('camera_clock')}")

    check("when the socket ends its clock leaves the status list",
          browser_clocks() == [], f"{browser_clocks()}")

    # Video C: the clock jumps backward mid-stream. The open event keeps
    # the clock it opened on and still closes in that domain; the
    # post-jump machine behaviour (invalid -> system) is section 2's.
    C0 = datetime(2026, 4, 2, 19, 0, 0)
    frames_c = [
        enveloped(jpeg(overlay(PEOPLE, C0 + timedelta(seconds=int(i * 0.4)))),
                  i * 0.4)
        for i in range(12)
    ] + [
        enveloped(jpeg(overlay(PEOPLE, datetime(2026, 4, 2, 18, 0, 0)
                               + timedelta(seconds=int(i * 0.4)))),
                  4.8 + i * 0.4)
        for i in range(8)
    ]
    asyncio.run(push(f"{WS_BASE}/api/restricted-zone/ws?overlay=json",
                     frames_c))

    jumped = zone_latest()
    jumped_closed = wait_closed(jumped["id"])
    jumped_lasted = (
        (datetime.fromisoformat(jumped_closed["ended_at"])
         - datetime.fromisoformat(jumped_closed["occurred_at"]))
        .total_seconds()
        if jumped_closed else None
    )

    # The opening moment is an estimate mapped from the wall onto the
    # recording's clock at lock — good to a couple of seconds, the
    # combined slack of one-second display granularity and the wall
    # drifting from the pushed positions during the hunt.
    jump_error = abs(
        (datetime.fromisoformat(jumped["occurred_at"]).replace(tzinfo=None)
         - C0).total_seconds()
    )
    check("a backward-jumping clock cannot drag its open event with it — "
          "the event keeps the clock it opened on and closes sanely",
          jump_error <= 3.0
          and jumped_closed is not None
          and 0.0 <= jumped_lasted <= 60.0,
          f"occurred {jumped['occurred_at']} ({jump_error:.0f}s from the "
          f"burned 19:00:00), lasted {jumped_lasted}s")

    section("7 · photos — one look, its own clock when it shows one")

    photo_reply = requests.post(
        f"{BASE}/api/restricted-zone/photo",
        files={"file": ("burned.jpg", jpeg(overlay(PEOPLE, T0)),
                        "image/jpeg")},
        timeout=60,
    ).json()

    check("the burned photo was analysed",
          photo_reply.get("success") is True, f"{photo_reply}")

    photo_event = zone_latest()
    check("a photo with a burned clock is stamped by it",
          photo_event["details"].get("timestamp_source") == "cctv"
          and photo_event["occurred_at"].startswith("2026-04-02T18:54:29"),
          f"{photo_event['occurred_at']} / {photo_event['details']}")

    plain_photo = requests.post(
        f"{BASE}/api/mask/photo",
        files={"file": ("plain.jpg",
                        (HERE / "fixtures" / "check_photo.jpg").read_bytes(),
                        "image/jpeg")},
        timeout=60,
    ).json()

    mask_events = requests.get(
        f"{BASE}/api/events",
        params={"module": "mask", "days": 1, "limit": 3},
        timeout=10,
    ).json()["data"]["events"]

    check("a photo with no burned clock is stamped by the system",
          plain_photo.get("success") is True
          and (not mask_events
               or mask_events[0]["details"].get("timestamp_source")
               == "system"),
          f"{mask_events[:1]}")

    section("6 · the server path — an uploaded recording, read by the "
            "server's own capture")

    video_path = WORKDIR / "burned_server.mp4"
    writer = cv2.VideoWriter(
        str(video_path), cv2.VideoWriter_fourcc(*"mp4v"), 12, (640, 480)
    )
    for i in range(12 * 40):
        writer.write(overlay(PEOPLE, T0 + timedelta(seconds=i // 12)))
    writer.release()

    upload = requests.post(
        f"{BASE}/camera/upload",
        files={"file": ("burned_server.mp4", video_path.read_bytes(),
                        "video/mp4")},
        timeout=60,
    ).json()
    check("the recording uploads", upload.get("success") is True,
          f"{upload}")

    sourced = requests.post(
        f"{BASE}/camera/source",
        json={"source": f"storage/uploads/{upload.get('filename')}"},
        timeout=30,
    ).json()
    check("and the server capture starts on it",
          sourced.get("success") is True, f"{sourced}")

    # The zone belongs to the camera it is drawn on: mark it again now
    # that the current source is the uploaded file.
    requests.post(
        f"{BASE}/api/restricted-zone/config",
        json={"polygon": ZONE, "frame_width": 640, "frame_height": 480},
        timeout=10,
    )

    # Watching the stream is what runs the analysis loop.
    stream = requests.get(
        f"{BASE}/api/restricted-zone/stream", stream=True, timeout=10
    )
    consumed_until = _time.monotonic() + 14.0
    try:
        for _chunk in stream.iter_content(chunk_size=8192):
            if _time.monotonic() > consumed_until:
                break
    finally:
        stream.close()

    status = requests.get(
        f"{BASE}/api/timestamp-clock/status", timeout=10
    ).json()
    clocks = status.get("data", {}).get("clocks", [])
    check("the source's clock reports itself locked while the recording "
          "plays",
          any(c.get("state") in ("locked", "coasting") for c in clocks),
          f"{clocks}")

    server_event = zone_latest()
    check("the server-path event is stamped by the recording's clock, "
          "anchored on the capture's own position",
          server_event["details"].get("timestamp_source") == "cctv"
          and server_event["occurred_at"].startswith("2026-04-02T18:5"),
          f"{server_event['occurred_at']} / "
          f"{server_event['details'].get('cctv_timestamp')}")

    section("8 · the surfaces — history, analytics, exports")

    summary = requests.get(
        f"{BASE}/api/events/summary", params={"days": 365}, timeout=10
    ).json()["data"]
    check("the day-by-day analytics bucket the April events under April",
          any(day.get("day") == "2026-04-02" for day in summary["by_day"]),
          f"{[day['day'] for day in summary['by_day']][:8]}")

    check("a year-wide window reaches the April events at all — the page "
          "offers exactly this window, and the events it was built for are "
          "older than every other one",
          any(row["occurred_at"].startswith("2026-04")
              for row in requests.get(
                  f"{BASE}/api/events",
                  params={"module": "restricted-zone", "days": 365,
                          "limit": 200},
                  timeout=10,
              ).json()["data"]["events"]),
          "no April row inside a 365-day window")

    # The export cap is a number the pages promise something about, so it
    # is served rather than copied: a frontend holding its own 500 would
    # keep promising it long after this constant moved.
    from app.api.event_routes import EXPORT_MAX_ROWS  # noqa: E402

    listing = requests.get(
        f"{BASE}/api/events", params={"limit": 1}, timeout=10
    ).json()["data"]

    check("both payloads carry the export cap the exports actually apply",
          listing.get("export_limit") == EXPORT_MAX_ROWS
          and summary.get("export_limit") == EXPORT_MAX_ROWS,
          f"list {listing.get('export_limit')}, summary "
          f"{summary.get('export_limit')}, server {EXPORT_MAX_ROWS}")

    csv_body = requests.get(
        f"{BASE}/api/events/export.csv", params={"days": 365}, timeout=30
    ).text
    csv_lines = [line for line in csv_body.splitlines() if line.strip()]
    check("the CSV export still opens with the pinned header, now with "
          "the clock column at its tail",
          csv_lines[0].startswith("When (UTC)")
          and csv_lines[0].endswith("Time source"),
          f"{csv_lines[0]}")

    check("and carries cctv rows at the recording's own time",
          any("2026-04-02" in line and line.endswith("cctv")
              for line in csv_lines[1:]),
          f"{[l for l in csv_lines[1:] if '2026-04-02' in l][:2]}")

    xlsx_bytes = requests.get(
        f"{BASE}/api/events/export.xlsx", params={"days": 365}, timeout=30
    ).content
    from openpyxl import load_workbook  # noqa: E402

    book = load_workbook(_io.BytesIO(xlsx_bytes))
    sheet = book.active
    headers = [cell.value for cell in sheet[1]]
    # Every row, not the first handful: the workbook is ordered newest
    # first, so rows stamped from footage — the ones this section is about
    # — sit at the bottom. The sheet is bounded by the export cap, so
    # reading all of it costs nothing.
    tails = [sheet.cell(row=r, column=9).value
             for r in range(2, sheet.max_row + 1)]
    whens = [sheet.cell(row=r, column=1).value
             for r in range(2, sheet.max_row + 1)]

    check("the Excel export carries nine columns ending in the clock "
          "column",
          headers[-1] == "Time source" and len(headers) == 9,
          f"{headers}")

    check("with real datetimes in When and a cctv row at the recording's "
          "time",
          any(str(when).startswith("2026-04-02") and tail == "cctv"
              for when, tail in zip(whens, tails)),
          f"{list(zip(whens, tails))[:3]}")

    # ------------------------------------------------------------------
    # Teardown — a verification suite must not leave its test polygons,
    # marks or cameras in the product's config. The zone marked while the
    # server capture ran clears against that source; stopping the capture
    # then lets the browser-bucket zone clear too; the phase suites that
    # may run after this one demand a clean start and are entitled to it.
    #
    # Registered to run on the way out as well as here, because the way
    # out is not always this line: a missing dependency once took this
    # suite down three checks earlier, the marked zone survived, and the
    # next phase suite spent seventeen red checks reporting a floor
    # somebody else had drawn. Cleaning up must not be contingent on
    # finishing.
    # ------------------------------------------------------------------
    _clean_up()
    leftovers = requests.get(
        f"{BASE}/api/restricted-zone/config", timeout=10
    ).json()
    check("the suite leaves no zone behind it",
          not (leftovers.get("data") or leftovers).get("polygon"),
          f"{leftovers}")


# ----------------------------------------------------------------------
# Verdict
# ----------------------------------------------------------------------

print()
if failures:
    print("FAILED:")
    for name in failures:
        print(f"  · {name}")
    print()
    print("The timestamp resolver does not ship.")
    sys.exit(1)

for name in advisories:
    print(f"  · {name}")
print("Every timestamp-resolver criterion measured so far holds.")
