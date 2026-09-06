"""
Safety event history and reporting.

    GET    /api/events                  the history, filtered and paged
    GET    /api/events/summary          the figures the reports page is built on
    GET    /api/events/export.csv       the same, as plain rows for any tool
    GET    /api/events/export.xlsx      the same, as an Excel workbook
    GET    /api/events/{id}             one event
    POST   /api/events/{id}/acknowledge sign it off with a conclusion
    GET    /api/events/{id}/snapshot    the picture that proves it

Order matters below: `/summary` and the two `/export.*` routes are declared
before `/{event_id}`, or FastAPI would match them as an event with the id
"summary".
"""

import csv
import io
from datetime import datetime, timedelta, timezone
from typing import Any, Optional

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field

from app.events.store import DISPOSITIONS, EVIDENCE_DIR, event_store
from app.modules import registry

router = APIRouter(prefix="/api/events", tags=["Safety events"])

#: Longest reporting window accepted, in days. A year of a busy site is a lot
#: of rows to aggregate on request; beyond this it wants a scheduled report,
#: not a page load.
MAX_PERIOD_DAYS = 365

#: Most events either export will write. A spreadsheet is for reading, and a
#: request that would build one out of tens of thousands of rows is a report
#: somebody should schedule instead.
#:
#: The number is served to the pages beside every total (`export_limit`)
#: rather than being written down again in the frontend, so what the screen
#: promises about an export cannot drift from what the export does. Both
#: pages say when a filter matches more than this — a file that silently
#: stops at its newest five hundred rows is the kind of quiet lie the rest
#: of this product spends its time removing.
EXPORT_MAX_ROWS = 500


def _since(days: int) -> str:
    """Start of the reporting window, as the stored timestamps are written."""
    days = max(1, min(int(days), MAX_PERIOD_DAYS))
    start = datetime.now(timezone.utc) - timedelta(days=days)
    return start.isoformat(timespec="seconds")


def _csv_safe(value: Any) -> str:
    """
    A cell that a spreadsheet will treat as text, not as a formula.

    Event summaries now carry names an operator typed — door names, notes —
    and a spreadsheet reads a cell beginning with =, +, - or @ as a formula to
    execute when the file is opened. Prefixing an apostrophe is the standard
    defence: Excel and LibreOffice both strip it on display, so the cell reads
    as written while never being run.
    """
    text = "" if value is None else str(value)

    return f"'{text}" if text[:1] in ("=", "+", "-", "@") else text


def _module_names() -> dict[str, str]:
    """Module ids to operator-facing names, so the UI shows neither raw ids
    nor its own copy of the list."""
    return {service.module_id: service.name for service in registry.list_services()}


class Acknowledgement(BaseModel):
    """What a human concluded about an event."""

    disposition: str = Field(
        description="valid, false_alarm or resolved",
    )
    note: Optional[str] = Field(
        default=None,
        max_length=1000,
        description="Optional free text, e.g. what was done about it.",
    )


@router.get("")
def list_events(
    module: Optional[str] = None,
    severity: Optional[str] = None,
    acknowledged: Optional[bool] = None,
    days: Optional[int] = Query(default=None, ge=1, le=MAX_PERIOD_DAYS),
    limit: int = Query(default=50, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
) -> dict[str, Any]:
    """
    The history, newest first.

    Every filter is optional and they combine, so the same endpoint answers
    "what is outstanding" and "everything the door camera saw last week".
    """
    page = event_store.list(
        module_id=module,
        severity=severity,
        acknowledged=acknowledged,
        since=_since(days) if days else None,
        limit=limit,
        offset=offset,
    )

    return {
        "success": True,
        # `export_limit` travels beside `total` so a page can tell whether an
        # export of this same filter would be whole, without keeping its own
        # copy of a number only this module decides.
        "data": {
            **page,
            "modules": _module_names(),
            "export_limit": EXPORT_MAX_ROWS,
        },
    }


@router.get("/summary")
def event_summary(
    days: int = Query(default=7, ge=1, le=MAX_PERIOD_DAYS),
) -> dict[str, Any]:
    """
    The figures behind the reports page.

    Aggregated in the database rather than by counting a list of events in the
    browser, so the page stays fast as the history grows and cannot disagree
    with the export.
    """
    summary = event_store.summary(_since(days))

    return {
        "success": True,
        "data": {
            **summary,
            "days": days,
            "modules": _module_names(),
            "export_limit": EXPORT_MAX_ROWS,
        },
    }


@router.get("/export.csv")
def export_events(
    days: int = Query(default=7, ge=1, le=MAX_PERIOD_DAYS),
    module: Optional[str] = None,
    severity: Optional[str] = None,
    acknowledged: Optional[bool] = None,
) -> StreamingResponse:
    """
    The filtered history as a spreadsheet.

    The same filters as the list, so what is exported is what was on screen —
    an export that quietly covers a different set than the page it came from
    is how reported numbers stop matching.
    """
    names = _module_names()

    page = event_store.list(
        module_id=module,
        severity=severity,
        acknowledged=acknowledged,
        since=_since(days),
        limit=EXPORT_MAX_ROWS,
    )

    buffer = io.StringIO()
    writer = csv.writer(buffer)

    # "Time source" rides at the tail so every reader of the original eight
    # columns keeps its positions: cctv when the moment came from the clock
    # burned into the footage, system when the server stamped it.
    writer.writerow(
        [
            "When (UTC)",
            "What",
            "Detail",
            "Severity",
            "Ended",
            "Signed off",
            "Conclusion",
            "Note",
            "Time source",
        ]
    )

    for event in page["events"]:
        writer.writerow(
            [
                _csv_safe(event["occurred_at"]),
                _csv_safe(names.get(event["module_id"], event["module_id"])),
                _csv_safe(event["summary"]),
                _csv_safe(event["severity"]),
                _csv_safe(event["ended_at"] or "still open"),
                _csv_safe(event["acknowledged_at"] or "no"),
                _csv_safe(event["disposition"] or ""),
                _csv_safe(event["note"] or ""),
                _csv_safe(
                    (event.get("details") or {}).get("timestamp_source")
                    or "system"
                ),
            ]
        )

    buffer.seek(0)
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="safety-events-{stamp}.csv"'
        },
    )


@router.get("/export.xlsx")
def export_events_xlsx(
    days: int = Query(default=7, ge=1, le=MAX_PERIOD_DAYS),
    module: Optional[str] = None,
    severity: Optional[str] = None,
    acknowledged: Optional[bool] = None,
) -> StreamingResponse:
    """
    The filtered history as an Excel workbook.

    Same rows and same filters as the CSV, so the two exports can never
    disagree; this one opens directly in Excel with types intact — the
    timestamp column is real datetimes Excel can sort and pivot, not text.

    One worksheet, one row per event, eight columns:

        When (UTC) · What · Detail · Severity · Ended · Signed off ·
        Conclusion · Note

    "What" is the capability that raised it, in the words the sidebar uses.
    "Ended" distinguishes a finished situation from one still standing when
    the export was taken. "Signed off / Conclusion / Note" carry the human
    review from Safety Events — real problem or false alarm, and anything
    the reviewer wrote.

    A cell here is typed data, never a formula: openpyxl writes strings as
    literal strings, so the =SUM() someone types into a door name arrives in
    Excel as text. The CSV needs an apostrophe defence for the same case;
    this format is safe by construction.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    names = _module_names()

    page = event_store.list(
        module_id=module,
        severity=severity,
        acknowledged=acknowledged,
        since=_since(days),
        limit=EXPORT_MAX_ROWS,
    )

    book = Workbook()
    sheet = book.active
    sheet.title = "Safety events"

    columns = [
        ("When (UTC)", 20),
        ("What", 22),
        ("Detail", 46),
        ("Severity", 10),
        ("Ended", 20),
        ("Signed off", 20),
        ("Conclusion", 12),
        ("Note", 32),
        ("Time source", 12),
    ]

    header_fill = PatternFill("solid", fgColor="16294A")
    header_font = Font(color="FFFFFF", bold=True)

    for index, (title, width) in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=index, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"

    def moment(value):
        """Stored ISO text as a real Excel datetime, or the text as-is."""
        if not value:
            return None
        try:
            stamp = datetime.fromisoformat(str(value))
            # Excel has no timezone; the column header says UTC instead.
            return stamp.replace(tzinfo=None)
        except ValueError:
            return str(value)

    for row, event in enumerate(page["events"], start=2):
        sheet.cell(row=row, column=1, value=moment(event["occurred_at"]))
        sheet.cell(row=row, column=2,
                   value=names.get(event["module_id"], event["module_id"]))
        sheet.cell(row=row, column=3, value=event["summary"])
        sheet.cell(row=row, column=4, value=event["severity"])
        sheet.cell(row=row, column=5,
                   value=moment(event["ended_at"]) or "still open")
        sheet.cell(row=row, column=6,
                   value=moment(event["acknowledged_at"]) or "no")
        sheet.cell(row=row, column=7, value=event["disposition"] or "")
        sheet.cell(row=row, column=8, value=event["note"] or "")
        sheet.cell(
            row=row, column=9,
            value=(event.get("details") or {}).get("timestamp_source")
            or "system",
        )

    for column in (1, 5, 6):
        letter = get_column_letter(column)
        for cell in sheet[letter][1:]:
            if cell.value is not None and not isinstance(cell.value, str):
                cell.number_format = "yyyy-mm-dd hh:mm:ss"

    payload = io.BytesIO()
    book.save(payload)
    payload.seek(0)

    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    return StreamingResponse(
        payload,
        media_type=(
            "application/vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
                f'attachment; filename="safety-events-{stamp}.xlsx"'
        },
    )


@router.get("/{event_id}")
def get_event(event_id: int) -> dict[str, Any]:
    event = event_store.get(event_id)

    if event is None:
        raise HTTPException(status_code=404, detail="No such event.")

    return {"success": True, "data": event}


@router.post("/{event_id}/acknowledge")
def acknowledge_event(event_id: int, body: Acknowledgement) -> dict[str, Any]:
    """
    Sign an event off with a conclusion.

    The conclusion is the valuable part. An operator marking false alarms is
    measuring the system, and that measurement is what the accuracy figure in
    the report is built from — without it, a count of events says nothing
    about whether the system is worth trusting.
    """
    try:
        event = event_store.acknowledge(event_id, body.disposition, body.note)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    if event is None:
        raise HTTPException(status_code=404, detail="No such event.")

    return {
        "success": True,
        "message": "Event signed off.",
        "data": event,
    }


@router.get("/{event_id}/snapshot")
def event_snapshot(event_id: int) -> FileResponse:
    """The picture taken when the event opened."""
    event = event_store.get(event_id)

    if event is None:
        raise HTTPException(status_code=404, detail="No such event.")

    name = event.get("snapshot")

    if not name:
        raise HTTPException(
            status_code=404, detail="No picture was saved for this event."
        )

    # Resolved and checked rather than joined blindly: the filename comes from
    # the database, and a path that escapes the evidence folder would serve
    # arbitrary files off the disk.
    path = (EVIDENCE_DIR / name).resolve()

    if not path.is_file() or not path.is_relative_to(EVIDENCE_DIR.resolve()):
        raise HTTPException(status_code=404, detail="The picture is missing.")

    return FileResponse(path, media_type="image/jpeg")


__all__ = ["router", "DISPOSITIONS"]
